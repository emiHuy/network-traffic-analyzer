"""
services/export.py
───────────────────────────────────────────────────────────────────────────────
Data export service layer.

Responsibilities
────────────────
• export_csv(session_id)
    Export captured packet data for a session as a CSV file.

• export_excel(session_id)
    Export captured packet data for a session as a styled Excel (.xlsx) file.

Both functions:
    - Validate session existence
    - Retrieve all packets for the session
    - Serialize them into the requested format
    - Return (file_bytes, filename)
───────────────────────────────────────────────────────────────────────────────
"""

import csv
import io
from db.packets import get_packets
from db.sessions import get_session

# Ordered column headers for both CSV and Excel exports
_HEADERS = [
    'session_id', 
    'session_name', 
    'src_ip', 
    'dst_ip', 
    'protocol', 
    'dst_port', 
    'size', 
    'timestamp'
]


def export_csv(session_id: int) -> tuple[bytes, str]:
    """
    Export session packet data as a CSV file.
    Returns the CSV file content as (bytes, filename).

    Raises
    ──────
    ValueError
        If the session does not exist.
    """
    session = get_session(session_id)
    if not session:
        raise ValueError(f'Session {session_id} not found.')
    
    packets = get_packets(session_id, limit=None, desc=False)
    
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_HEADERS)
    writer.writeheader()
    
    # Write packet data
    for pkt in packets:
        writer.writerow({
            'session_id': session['id'],
            'session_name': session['name'],
            **pkt,
        })

    # Generate safe file name
    safe_name = session['name'].replace(' ', '_')
    filename = f'session_{session_id}_{safe_name}.csv'

    return buf.getvalue().encode('utf-8'), filename


def export_excel(session_id: int) -> tuple[bytes, str]:
    """
    Export session packet data as a formatted Excel (.xlsx) file.
    Returns the Excel file content as (bytes, filename).

    Raises
    ──────
    RuntimeError
        If openpyxl is not installed.
    ValueError
        If the session does not exist.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError('openpyxl is required: pip install openpyxl')

    session = get_session(session_id)
    if session is None:
        raise ValueError(f'Session {session_id} not found')

    packets = get_packets(session_id, limit=None, desc=False)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Session {session_id}'

    # Header styling
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', name='Calibri')

    # Write header row
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Populate rows with packet data
    for row_idx, pkt in enumerate(packets, start=2):
        ws.cell(row=row_idx, column=1, value=session['id'])
        ws.cell(row=row_idx, column=2, value=session['name'])
        ws.cell(row=row_idx, column=3, value=pkt['src_ip'])
        ws.cell(row=row_idx, column=4, value=pkt['dst_ip'])
        ws.cell(row=row_idx, column=5, value=pkt['protocol'])
        ws.cell(row=row_idx, column=6, value=pkt['dst_port'])
        ws.cell(row=row_idx, column=7, value=pkt['size'])
        ws.cell(row=row_idx, column=8, value=pkt['timestamp'])

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Save workbook to in-memory bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Generate safe filename
    safe_name = session['name'].replace(' ', '_')
    filename = f"session_{session_id}_{safe_name}.xlsx"

    return buf.read(), filename